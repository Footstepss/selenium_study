import pandas as pd
from openpyxl  import load_workbook

class DataFrame(object):
	
	def __init__(self,casepath,sheet_name):
		self.testcase = pd.read_excel(casepath,sheet_name = sheet_name)

	def fill_null(self,column):
		#处理excel中存在合并的单元格导致dataframe为空的的数据项,method = 'ffill'向上填充
		self.testcase[column].fullna(method='ffill')

	def write_excel(self,casepath,sheet_name):
		#因为直接to_excel会覆盖原来的数据 需要引入openpyxl库
		book = load_workbook(casepath)
		excelWriter = pd.ExcelWriter(casepath,engine='openpyxl')
		excelWriter.book = book
		self.testcase.to_excel(excelWriter,sheet_name=sheet_name)
		excelWriter.close()
		#self.testcase.to_excel(casepath,sheet_name = sheet_name)

	def get_testcase(self):
		return self.testcase

if __name__ == '__main__':
	test = DataFrame('C:\\Users\\Z8647\\Desktop\\data_test02.xlsx','Sheet5')
	print(test)
	book = load_workbook('C:\\Users\\Z8647\\Desktop\\data_test02.xlsx')
	excelWriter =pd.ExcelWriter('C:\\Users\\Z8647\\Desktop\\data_test02.xlsx',engine='openpyxl')
	print(excelWriter.book)
	print(book)
	excelWriter.book = book
	test.testcase.to_excel(excelWriter,sheet_name='Sheet6')
	excelWriter.close()
	